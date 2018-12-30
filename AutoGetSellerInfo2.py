# -*- coding:UTF-8 -*-


from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By



import os
import re
import datetime as dt
import sys
import time
from openpyxl import load_workbook
from selenium import webdriver
from sys import stdin

DATA_DIR = os.path.abspath(os.path.dirname(__file__)) + "\\data\\ProductData.xlsx"
SELLER_HREF = "https://www.amazon.com/sp?_encoding=UTF8&seller="
ASIN_HREF = "https://www.amazon.com/dp/"
    
def isAmzonCheck(driver):
    try:
        e_h4 = driver.find_elements_by_xpath("//h4[text()='Type the characters you see in this image:']")
        if len(e_h4) > 0:
            print("////////////////////////////////////")
            print("遇到Amazon验证，请手动验证，再按‘回车’继续...")
            stdin.readline()
        else:
            return True
    except:
        return True
            
def ifUpdataAsin():
    print("////////////////////////////////////")
    print('是否更新所有买家的asin？（不更新请输入N，否则直接回车继续）：')
    n = sys.stdin.readline().strip('\n')
    if n == 'N':
        return False
    else:
        return True

def LoadProductData(worksheet):
    sellers = {}
    sellerDatas = []
    
    col = worksheet.max_column
    row = worksheet.max_row
    
    if col == 1:
        print('表格无seller数据，请确认输入后再运行！')
        return sellers
        
    print('正在获取卖家信息，请稍后。')
    for c in range(2, row + 1):
        name = worksheet.cell(row = 1,column = c).value
        seller = worksheet.cell(row = 2,column = c).value
        if name == None or seller == None:
            break
        sellerDatas.append(seller)
        for r in range(3,row):
            cValue =  worksheet.cell(row = r,column = c).value
            if cValue != None:
                sellerDatas.append(cValue)
        sellers[name] = sellerDatas
        sellerDatas = []
        
    print('已获得以下卖家信息：')
    for key in sellers:
        s = sellers[key]
        print('sellerID of %s ： %s'%(key,s[0]))
        for item in range(1,len(s)):
            print(s[item])
        
    return sellers

def GetAsinsReview(driver,sellerDatas):
    asins = []
    asinRev = {}
    
    re_asin = r'/dp/(.*)\?m='
    #打开卖家链接
    driver.get(SELLER_HREF + sellerDatas[0])
    #等待加载后点击product页面
    isAmzonCheck(driver)
    e_lis = driver.find_elements_by_xpath("//li[contains(@class,'a-tab-heading')]")
    e_lis[len(e_lis) - 1].click()
    time.sleep(4)
    
    #等待加载后获取页数
    pages = 1
    e_pages = driver.find_elements_by_xpath("//ul[contains(@class,'a-pagination')]//li[contains(@class,'products-pagination-button')]")
    if e_pages != []:
        e_page = e_pages[len(e_pages)-2]
        pages = int(e_page.get_attribute('id'))
    print("产品页数：%d"%(pages))

    #循环抓取产品asin跟review数量，并点击next直到不可点击
    for i in range(1, (pages + 1)):
        print("获取第 %d 页产品asin"%(i))
        #获取产品元素块
        e_divs = driver.find_elements_by_xpath("//div[contains(@class,'a-column product-column a-span2')]//div[contains(@class,'a-section product-details')]")
        for e_div in e_divs:
            e_a = e_div.find_element_by_class_name('a-link-normal')
            asinStr = re.findall(re_asin, e_a.get_attribute('href'))
            try:
                e_rating = e_div.find_element_by_class_name('product-rating')
                reviewNum = e_rating.text
            except:
                reviewNum = 0
            print("获得asin：%s ,reveiw数量为：%s"%(asinStr, reviewNum))
            asins.append(asinStr[0])
            asinRev[asinStr[0]] = reviewNum
        #翻页
        if i != pages:
            time.sleep(2)
            isAmzonCheck(driver)
            e_btns = driver.find_elements_by_xpath("//ul[contains(@class,'a-pagination')]//li[contains(@class,'products-pagination-button')]")
            e_btn = e_btns[len(e_btns)-1]
            e_btn.click()
            time.sleep(4)
            #页面产生跳转，需要将handle定位到当前页面
            sreach_window = driver.current_window_handle
    
    print("asin总共有 %d 个"%(len(asins)))
    
    if len(sellerDatas) == 1:
        sellerDatas = sellerDatas + asins  
    else:
        sellerid = sellerDatas[0]
        del sellerDatas[0]
        sellerDatas = list(set(sellerDatas).union(set(asins)))
        sellerDatas.insert(0,sellerid)
    
    return sellerDatas, asinRev
    
#datas是list，[0]是sellerID，后面是asins
def updateAsins(worksheet,datas):
    colAsins = []
    rowLen = 0
    sellerid = datas[0]
    dlength = len(datas)
    sellerAsins = datas[1 : dlength]
    col = worksheet.max_column
    
    for c in range(2, col + 1):
        if sellerid == worksheet.cell(row = 2,column = c).value:
            for cell in list(worksheet.columns)[c - 1]:
                if cell.value is not None:
                    rowLen += 1
            for r in range(2, rowLen + 1):
                colAsins.append(worksheet.cell(row = r,column = c).value)
            for sellerAsin in sellerAsins:
                if sellerAsin not in colAsins:
                    rowLen += 1
                    worksheet.cell(row = rowLen,column = c,value = sellerAsin)
            break

    return  

def writeReviewData(workbook,sellerName,asinsRev):
    sheets = workbook.get_sheet_names()
    asins = asinsRev.keys()
    UpdateTime = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #工作薄已有asin
    wsAsins = []
    #asin的数量
    row = 0
    
    #直接判断sellerName是否包含在sheets中，不包含直接创建该工作薄
    if sellerName in sheets:
        print("%s 存在工作薄"%(sellerName))
        ind = sheets.index(sellerName)
        print("正打开 %s 工作薄"%(sellerName))
        wsheet = workbook.get_sheet_by_name(sheets[ind]) 
        maxRow = wsheet.max_row
        for r in range(2, maxRow + 1):
            wsAsins.append(wsheet.cell(row = r,column = 1).value)
        row = len(wsAsins)
        print("原有 %d 行数据，最后获得 %d 行数据"%(maxRow,row))
        for asin in asins:
            if asin not in wsAsins:
                row += 1
                wsheet.cell(row = (row + 1),column = 1).value = asin
                wsAsins.append(asin)
        print("现有 %d 行数据"%(row))
    else:
        print("%s 需创建工作薄"%(sellerName))
        wsheet = workbook.create_sheet(title = sellerName)
        wsheet.cell(row = 1,column = 1).value = 'asin'
        print(" %s 工作薄已创建"%(sellerName))
        for asin in asins:
            row += 1
            wsheet.cell(row = (row + 1),column = 1).value = asin
            wsAsins.append(asin)
    
    print("开始写入数据")
    col = wsheet.max_column
    wsheet.cell(row = 1,column = (col + 1)).value = UpdateTime 
    for key in asinsRev:
        asinIndex = wsAsins.index(key) + 1
        wsheet.cell(row = (asinIndex + 1),column = (col + 1)).value = asinsRev[key]
    
    print("数据写入完成")
    
    return

if __name__=="__main__":
    try:
        driver = None
        productAsins = {}
        
        #进入ProductData.xlsx，获取sellerInfo工作薄所有seller跟asin数据
        #判断是否需要更新seller的asin信息：
        #是 则遍历sellerid：
        #   进入seller页面
        #   进入product栏
        #   抓取总页数
        #   循环抓取产品asin，并点击next直到不可点击
        #   存入ProductData.xlsx
        #否 则直接进入下一步
        #遍历sellerid：
        #   遍历asin：
        #       进入asin页面
        #       抓取review数据
        #       （这里可以拓展为获取产品所有可用信息，包括价格，review，排名等）
        #   存入ProductData.xlsx（如果从文档中获得上一次的数据，可以考虑做一次计算新增值）
        #结束退出
        
        #从ProductData.xlsx加载产品列表文件，获取产品名及产品链接
        print("获取产品信息")
        #打开ProductData.xlsx
        wb = load_workbook(filename=DATA_DIR)
        sheets = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheets[0])
        
        #获取产品信息
        sellersDatas = LoadProductData(ws)
        
        #更新asin数据
        if len(sellersDatas) > 0:
            #打开Chrome
            options = webdriver.ChromeOptions()
            driver = webdriver.Chrome(chrome_options=options)
            for key in sellersDatas:
                sellerDatas = sellersDatas[key]
                print('开始获取 %s 最新的asin数据与review数据'%(key))
                #更新卖家asin
                (sellerDatas,asinRev) = GetAsinsReview(driver,sellerDatas)
                print("已获取完 %s 所有产品asin及review数量！"%(key))
                sellersDatas[key] = sellerDatas
                print("开始更新 %s asin信息！"%(key))
                updateAsins(ws,sellerDatas)
                print("开始更新 %s asin的数量！"%(key))
                writeReviewData(wb,key,asinRev)
                print("已更新完 %s asin的数量！"%(key))

            #把文件另存为到Data2.xlsx中，防止程序运行中数据文件损坏
            wb.save(os.path.abspath(os.path.dirname(__file__)) + "\\data\\result\\Data2.xlsx")
                    
            print('已更新完卖家信息！')
        
        else:
            print('data文件夹中ProductData.xlsx未有seller数据，请确认已填入数据并保存后再试')

            
        #结束退出
        print("程序结束！")
        
        
    finally:
        if driver is not None:
            driver.close()    
